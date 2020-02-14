from openpyxl import Workbook
from openpyxl import load_workbook
import urllib3
import json
import os
os.chdir(r'e:\code\叮咚商城')
excl="1111.xlsx"
wb = load_workbook(excl)
wb.guess_types = True 
ws=wb.active

http = urllib3.PoolManager()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
logindata = {"isRememberPwd":1,"password":"666666","userName":"77019900001"}
encoded_login = json.dumps(logindata).encode('utf-8')

# print(encoded_shop)
rlogin = http.request(
     'POST',
     'https://s.197.com/web-smp/user/login/in',
     body=encoded_login,
     headers={'Content-Type': 'application/json','User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
     )


httpdata=json.loads(rlogin.data.decode('utf-8'))
token=httpdata["data"]
#添加cookie
# print(token["token"])
# Authorization=gwzkvH0522JVYKUDMV
cookie ="Authorization" + "=" + token["token"]
print(cookie)
if httpdata["code"] == "000000":

    print("登陆成功：")
    for rannum in range(ws.max_row):  
	    rannum+=1
		#B名称
	    b = "B" + str(rannum) 
		#E价格
	    e = "E" + str(rannum) 
		#E图片
	    f = "F" + str(rannum) 
		#分类id
	    flid=403502
	    imgurl=ws[f].value
	    shopname=ws[b].value
	    price=ws[e].value
	    shopdata = {
			"enableSupperMemberPrice": 1,
			"categoryIds": [flid],
			"coverImage": imgurl,
			"imgTextHybr": [{
				"img": imgurl
			}],
			"goodsDesc": "",
			"imgUrls": [],
			"mainImagesUrl": [imgurl],
			"inventory": 99999,
			"isInfiniteInventory": 2,
			"isSpecial": 1,
			"name": shopname,
			"originalPrice": 0,
			"price": price,
			"type": 1,
			"tagType": "null",
			"moneyType": 1,
			"isSupportMemberCardPay": 2,
			"openProductDistributionSetting": 1,
			"restrictedPurchase": 1,
			"status": 1,
			"openLikeAndComment": 2
		}
		#转义
	    encoded_shop = json.dumps(shopdata).encode('utf-8')
	    radd = http.request(
		'POST',
		'https://s.197.com/web-smp/smp/org/product/add',
		body=encoded_shop,
		headers={
		'cache-control': "no-cache",
		'Content-Type': 'application/json',
		"Cookie": cookie,
		'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
		}
		)
	    addmsg = httpdata=json.loads(radd.data.decode('utf-8'))
	    print(str(addmsg["code"]) +addmsg["msg"]+ "操作时间戳:" + str(addmsg["currentTimes"]))    
else:
    print("登陆失败。请联系管理员")