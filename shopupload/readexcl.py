from openpyxl import Workbook
from openpyxl import load_workbook
import os
import urllib3
import json
http = urllib3.PoolManager()
os.chdir(r'e:\code\叮咚商城')

excl="1111.xlsx"
url = 'https://s.197.com/web-smp/smp/snsImgUpload'
wb = load_workbook(excl)
wb.guess_types = True 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

ws=wb.active
print(ws.max_row)
#拼接循环
for rannum in range(ws.max_row):  
    rannum+=1
    a = "A" + str(rannum)   
    # 拼接图片
    imgcode = str(ws[a].value)+".jpg"
    print("图片名称:"+imgcode)
    print ("目前行数:" + str(rannum))
    #读取文件发送请求
    with open(imgcode,'rb') as fp:
    #  print(fp)
     file_data = fp.read()
     ##接口不需要在apy登陆
    r = http.request(
   'POST',
   'https://s.197.com/web-smp/smp/snsImgUpload',
    headers={
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'
     },    
     fields={
        'filefield': (imgcode, file_data),
     }
     )
    f = "F" + str(rannum)
    httpdata=json.loads(r.data.decode('utf-8'))
    if httpdata["code"] == "000000":
        imgroute = str(httpdata["data"]).split("|")[0]
#截取前19位后函数
        ws[f]= imgroute[19:]
        print("目前图片路径:" +imgroute[19:])
    else:
        ws[f]= "fail"
    # ws[f]= str(r.text)
wb.save(excl)
